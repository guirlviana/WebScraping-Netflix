from selenium import webdriver
import os
import openpyxl
import random
import time
from selenium.webdriver.chrome.options import Options
class PopularesNetflix:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--disable-gpu')
        self.driver = webdriver.Chrome(executable_path= os.getcwd() + os.sep + 'chromedriver.exe', options=chrome_options)

    def Iniciar(self):
        self.driver.get('https://www.netflix.com/br/browse/genre/839338')
        self.titulos = []   
        time.sleep(2)
        self.FazerPlanilha()
        self.EncontrarElementos()
        self.sugestao()
    def FazerPlanilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('Catalogos Populares Netflix Julho')
        self.planilha_valores = self.planilha['Catalogos Populares Netflix Julho']
        self.planilha_valores.cell(row= 1, column=1, value='Titulos')
        self.planilha_valores.cell(row= 1, column=3, value='Sugestão')
        
    def EncontrarElementos(self):
        try: 
            for self.cont in range(0, 11):
                self.titulo = self.driver.find_elements_by_xpath('//span[@class="nm-collections-title-name"]')
                self.ArmazenarValores()  
                     
        except Exception as erro:
            print('Fim')
    
    def ArmazenarValores(self):
        nova_linha = [self.titulo[self.cont].text] 
        self.planilha_valores.append(nova_linha)
        self.titulos.append(str(nova_linha))
        
    
    def sugestao(self):    
        self.sugestao = random.choice(self.titulos).replace("'", "").replace("[", "").replace("]", "")
        self.planilha_valores.cell(row=2, column=3, value= self.sugestao)
        self.planilha.save(f'Populares netflix Julho.xlsx')

root = PopularesNetflix()
root.Iniciar()


